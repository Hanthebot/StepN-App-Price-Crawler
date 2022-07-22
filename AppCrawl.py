import json
import os
import socket
import sys
import time
import traceback
import winsound

import colorama
import cv2
import numpy as np
from openpyxl import Workbook
from PIL import Image
import pyautogui
import pytesseract
import requests
import telepot
from termcolor import colored

from util import fl, fo2, fo3, fo4, mkIfNone, number, onlyNumber, randString

pytesseract.pytesseract.tesseract_cmd = r"C:/Program Files/Tesseract-OCR/tesseract.exe"

setting = {
    "coordPrice_C": [0.0631, 0.5390, 0.2523, 0.0203],
    "retryDelay": 0.25,
    "refreshPoint_C": [0.4865, 0.2604],
    "dragLength_C": 0.1478,
    "dragDuration": 0.2,
    "dragDelay": 2,
    "detectionMaxTry": 10,
    "saveImg_C": [0.0541, 0.2401, 0.4505, 0.3445],
    "dimension": [1315, 33, 555, 987],
    "categories": ["Trainer"]
}

print("setting...", end = "")
if os.path.exists("./data/data.json"):
    with open("./data/data.json", "r", encoding = "utf-8") as fil:
        totalSet = json.load(fil)
        if socket.gethostname() in totalSet.keys():
            setting = totalSet[socket.gethostname()]
            print("found")
        else:
            setting = totalSet["default"]
            print("default")
else:
    print("default")

if "place" in setting["optional"]:
    try:
        cmdXi, cmdYi, cmdX, cmdY = setting["optional"]["place"]
        os.system(f"cmdow @ /MOV {cmdXi} {cmdYi} /SIZ {cmdX} {cmdY}")
    except OSError:
        pass

def coorCalculate(ratio, type_):
    xi, yi, x, y = setting["dimension"]
    if type_ == "length_x":
        return int(ratio * x)
    elif type_ == "length_y":
        return int(ratio * y)
    elif type_ == "point_xy":
        return (int(xi + ratio[0] * x), int(yi + ratio[1] * y))
    elif type_ == "dimension_xy":
        return (int(xi + ratio[0] * x), int(yi + ratio[1] * y), int(ratio[2] * x), int(ratio[3] * y))

def setting_unpack():
    setting["dimension"] = tuple(setting["dimension"])
    setting["dragLength"] = coorCalculate(setting["dragLength_C"], "length_y")
    for key in setting["point_xy_C"]:
        setting[key] = coorCalculate(setting["point_xy_C"][key], "point_xy")
    setting["categoryCoor"] = {}
    for key in setting["categoryCoor_C"]:
        setting["categoryCoor"][key] = coorCalculate(setting["categoryCoor_C"][key], "point_xy")
    for key in setting["dimension_xy_C"]:
        setting[key] = coorCalculate(setting["dimension_xy_C"][key], "dimension_xy")

setting_unpack()

info = {
    "developer": "",
    "developer_id": "",
    "admin": [],
    "bot_token": "",
    "bot_id": "",
    "alert_bot_token": "",
    "subscribers": [],
    "whitelist": [],
    "subscriber_name_dict": {},
    "server_list": {},
    "formula": "(360 * {GST} + 40 * {GMT}) * 1.06",
    "interval": 0.5,
    "msgGap": 0.1,
    "errLim": 10,
    "msgTry": 10,
    "period": 60
}

print("ID information...", end = "")
if os.path.exists("./data/StepNinfo.json"):
    with open("./data/StepNinfo.json", "r", encoding = "utf-8") as fil:
        try:
            info = json.load(fil)
            print("found")
        except json.decoder.JSONDecodeError:
            print("default")
else:
    print("default")

bot = telepot.Bot(info["bot_token"])

alert_bot = telepot.Bot(info["alert_bot_token"])

user_data = {
    "default": {
        "ratioMax": 28,
        "ratioMin": 26,
        "percentage": {category: 0.0 for category in setting["categories"] + ["general"]},
        "priceMax": {category: 60 for category in setting["categories"] + ["general"]},
        "priceMin": {category: 20 for category in setting["categories"] + ["general"]}
    }
}

print("user data...", end = "")
if os.path.exists("./data/StepNuserData.json"):
    with open("./data/StepNuserData.json", "r", encoding = "utf-8") as fil:
        try:
            user_data = json.load(fil)
            print("found")
        except json.decoder.JSONDecodeError:
            bot.sendMessage(info["developer"], "ERROR")
            print("default")
else:
    print("default")

for you in info["subscribers"]:
    if you not in list(user_data.keys()):
        user_data[you] = user_data["default"]
        try:
            bot.sendMessage(you,"Your data was reset")
        except telepot.exception.TelepotException:
            pass

def saveFile():
    with open("./data/StepNuserData.json", "w", encoding = "utf-8") as fil_:
        json.dump(user_data, fil_, indent = 4)

def saveInfo():
    with open("./data/StepNinfo.json", "w", encoding = "utf-8") as fil_:
        json.dump(info, fil_, indent = 4, ensure_ascii = False)

manifest = {
    "name": "StepN",
    "version": "Jun27",
    "curB": "",
    "commands": {
        "퍼센트": "최저가 알림 체크/업데이트 \n예: 퍼센트/퍼센트 1.0",
        "ratiomax": "SOL/GST 비율 알림 체크/업데이트 \n예: ratiomax/ratiomax 27.0",
        "ratiomin": "SOL/GST 비율 알림 체크/업데이트 \n예: ratiomin/ratiomin 27.0",
        "pricemax": "StepN 가격 알림 체크/업데이트 \n예: pricemax [category/general]/pricemax [category/general] 60",
        "pricemin": "StepN 가격 알림 체크/업데이트 \n예: pricemin [category/general]/pricemin [category/general] 20",
        "데이터": "프로그램용 내부 크롤링 데이터 출력 \n예: 데이터",
        "유저데이터": "유저 개인 설정 \n예: 데이터",
        "기본데이터": "유저 개인 설정 \n예: 데이터",
        "기본프린트": "프로그램용 출력 \n예: 기본프린트",
        "프린트": "프로그램용 출력 \n예: 프린트",
        "초대/링크/초대링크": "프로그램 공유용 링크",
        "server": "returns server name",
        "help": "message list",
        "usage": "returns usage \n 예:usage/usage [command]",
        "formula": "returns cost calculation formula \n 예:formula"
    },
    "commands_admin": {
        "refresh_code": "",
        "code": "",
        "execute": "",
        "kill_execute": "",
        "us": "",
        "whole_data": "",
        "naked_data": "",
        "revive": "",
        "퍼센트d": "최저가 알림 체크/업데이트 \n예: 퍼센트d/퍼센트d 1.0",
        "ratiomaxd": "SOL/GST 비율 알림 체크/업데이트 \n예: ratiomaxd/ratiomaxd 27.0",
        "ratiomind": "SOL/GST 비율 알림 체크/업데이트 \n예: ratiomind/ratiomind 27.0",
        "pricemaxd": "StepN 가격 알림 체크/업데이트 \n예: pricemaxd [category/general]/pricemaxd [category/general] 60",
        "pricemind": "StepN 가격 알림 체크/업데이트 \n예: pricemind [category/general]/pricemind [category/general] 20",
        "die": "",
        "helpd": "",
        "usaged": "",
        "formula": "계산식 체크/변경 \n예: formula ",
        "screenshot": "스크린샷 전송. \n예: screenshot OR screenshot 130 400 100 100"
    }
}

data = {
    "market": {},
    "prev_market": {},
    "users": {you: {} for you in info["subscribers"] + ["default"]},
    "start": time.strftime("%Y%m%d%H%M%S"),
    "die": False,
    "revive": False,
    "error": "None",
    "invitation_code": randString(7),
    "execute_on": time.time(),
    "cycle_i": time.time(),
    "cycle_f": time.time(),
    "initial": True,
    "coinlist": ["SOL", "GST", "GMT"],
    "recent_ban_date": {you: "0" for you in info["subscribers"]},
    "alertOK": {you: False for you in info["subscribers"] + ["default"]}
}

mkIfNone("./dataxl")
mkIfNone("./img")
mkIfNone("./data")
mkIfNone("./log")
wb = Workbook()
ws1 = wb.active
colorama.init()

def upbit(BTC = "BTC", cur = "KRW"):
    htm = requests.get(f"https://api.upbit.com/v1/orderbook?markets={cur}-{BTC}", headers = {"User-Agent": "Mozilla/5.0"})
    hack = htm.json()
    
    return [hack[0]["orderbook_units"][0]["bid_price"], hack[0]["orderbook_units"][0]["ask_price"], hack[0]["orderbook_units"][0]["bid_size"], hack[0]["orderbook_units"][0]["ask_size"]]

def ftx(BTC = "BTC", cur = "USD"):
    htm = requests.get(f"https://ftx.com/api/markets/{BTC}/{cur}/orderbook", headers = {"User-Agent": "Mozilla/5.0"})
    hack = htm.json()
    
    return [hack["result"]["bids"][0][0], hack["result"]["asks"][0][0], hack["result"]["bids"][0][1], hack["result"]["asks"][0][1]]

def scanSection(cord, showImage = False):
    img = pyautogui.screenshot(region=cord)
    imageFiltered = Image.fromarray(cv2.cvtColor(np.array(img), cv2.COLOR_BGR2GRAY))
    if showImage:
        imageFiltered.show()
    stringified = pytesseract.image_to_string(imageFiltered)
    return stringified

def read_screen(coordinate, maxTry = setting["detectionMaxTry"]):
    lis = {"valid": False}
    score = 0
    while score < maxTry and not lis["valid"]:
        time.sleep(setting["retryDelay"])
        scanned = scanSection(coordinate)
        scanned = scanned.split(" ")[0]
        filtered = onlyNumber(scanned)
        filteredP = filtered.replace(".", "")
        if number(filtered) or number(filteredP):
            lis["value"] = float(filtered) if number(filtered) else float(filteredP)
            lis["valid"] = True
            break
        score += 1
    return lis

def refresh(coor1 = setting["refreshPoint"], drag = setting["dragLength"],):
    pos = pyautogui.position()
    pyautogui.moveTo(coor1)
    pyautogui.dragTo(coor1[0], coor1[1] + drag, setting["dragDuration"], button = "left")
    pyautogui.moveTo(pos)
    time.sleep(setting["dragDelay"])

def setCategory(category_, duration = setting["dragDelay"]):
    pos = pyautogui.position()
    pyautogui.click(setting["filter"])
    time.sleep(setting["clickDelay"])
    if setCategory.initial:
        setCategory.initial = False
        pyautogui.click(setting["clear"])
        time.sleep(duration)
        pyautogui.click(setting["filter"])
        time.sleep(setting["clickDelay"])
        pyautogui.click(setting["sneakers"])
        time.sleep(setting["clickDelay"])
    pyautogui.click(setting["categoryCoor"][category_])
    time.sleep(setting["clickDelay"])
    pyautogui.click(setting["confirm"])
    time.sleep(setting["clickDelay"])
    pyautogui.moveTo(pos)
    time.sleep(duration)

setCategory.initial = True

def stepNPrice(coordinate = setting["coordPrice"]):
    filtered_data = {}
    for category_ in setting["categories"]:
        setCategory(category_)
        filtered_data[category_] = read_screen(coordinate)
    return filtered_data

def getcoin():
    base = {BTC: [] for BTC in data["coinlist"]}
    for BTC in data["coinlist"]:
        err = 0
        prblm = True
        while err < info["errLim"] and prblm:
            try:
                base[BTC] = ftx(BTC, manifest["curB"]) if manifest["curB"] != "" else ftx(BTC)
                prblm = False
            except KeyError:
                err += 1
                time.sleep(info["interval"])
    shoesPrice = stepNPrice()
    
    data["market"] = {
        "base": {BTC: {} for BTC in data["coinlist"]},
        "shoes": {category_: {} for category_ in setting["categories"]}
    }
    try:
        for BTC in data["coinlist"]:
            data["market"]["base"][BTC]["bid"] = float(base[BTC][0])
            data["market"]["base"][BTC]["ask"] = float(base[BTC][1])
        data["market"]["cost"] = (360 * float(base["GST"][1]) + 40 * float(base["GMT"][1])) * 1.06
        data["market"]["SOLGSTRatio"] = float(base["SOL"][0]) / float(base["GST"][1])
        for category_ in setting["categories"]:
            if shoesPrice[category_]["valid"]:
                data["market"]["shoes"][category_]["price"] = shoesPrice[category_]["value"]
                data["market"]["shoes"][category_]["priceUSD"] = float(base["SOL"][0]) * shoesPrice[category_]["value"]
                data["market"]["shoes"][category_]["premium"] = (data["market"]["shoes"][category_]["priceUSD"] / data["market"]["cost"] - 1) * 100
            else:
                data["market"]["shoes"][category_] = data["prev_market"]["shoes"][category_]
            data["market"]["shoes"][category_]["valid"] = shoesPrice[category_]["valid"]
        for BTC in data["coinlist"]:
            n = data["coinlist"].index(BTC)
            ws1.cell(ws1.max_row, 2 * n + 2, float(base[BTC][0]))
            ws1.cell(ws1.max_row, 2 * n + 3, float(base[BTC][1]))
        ws1.cell(ws1.max_row, 8, data["market"]["cost"])
        ws1.cell(ws1.max_row, 9, data["market"]["SOLGSTRatio"])
        for category_ in setting["categories"]:
            n = setting["categories"].index(category_)
            ws1.cell(ws1.max_row, 3 * n + 10, data["market"]["shoes"][category_]["price"])
            ws1.cell(ws1.max_row, 3 * n + 11, data["market"]["shoes"][category_]["priceUSD"])
            ws1.cell(ws1.max_row, 3 * n + 12, data["market"]["shoes"][category_]["premium"])
    except KeyError:
        data["market"] = data["prev_data"]
    
def printerG(category_, you_ = "default", color = False):
    pre_val = fo2(data["market"]["shoes"][category_]["premium"])
    if color:
        pre_val = colored(pre_val, "green" if data["market"]["shoes"][category_]["premium"] > float(user_data[you_]["percentage"][category_]) else "red")
    instantS = ""
    instantS += f"StepN {fl(setting['categoriesName'], category_)}:  {data['market']['shoes'][category_]['price']} SOL" + "\n"
    instantS += f"달러 가격:  {fo2(data['market']['shoes'][category_]['priceUSD'])} USD" + "\n"
    instantS += f"프리미엄: {pre_val}%" + "\n"
    instantS += "\n"
    return instantS
    
def firstGS(you_ = "default", color = False):
    pre_val = max([data["market"]["shoes"][category_]["premium"] for category_ in setting["categories"]])
    catPre = fl(setting["categoriesName"], setting["categories"][[data["market"]["shoes"][category_]["premium"] for category_ in setting["categories"]].index(pre_val)])
    pre_col = fo2(pre_val)
    min_val = min([data["market"]["shoes"][category_]["priceUSD"] for category_ in setting["categories"]])
    catMin = fl(setting["categoriesName"], setting["categories"][[data["market"]["shoes"][category_]["priceUSD"] for category_ in setting["categories"]].index(min_val)])
    min_col = fo2(min_val)
    max_val = max([data["market"]["shoes"][category_]["priceUSD"] for category_ in setting["categories"]])
    catMax = fl(setting["categoriesName"], setting["categories"][[data["market"]["shoes"][category_]["priceUSD"] for category_ in setting["categories"]].index(max_val)])
    max_col = fo2(max_val)
    if color:
        pre_col = colored(pre_col, "green" if pre_val > user_data[you_]["percentage"]["general"] else "red")
        min_col = colored(min_col, "green" if min_val < user_data[you_]["priceMin"]["general"] else "red")
        max_col = colored(max_col, "green" if max_val > user_data[you_]["priceMax"]["general"] else "red")
    instantS = ""
    instantS += f"프리미엄 최대:  {pre_col}% -{catPre}" + "\n"
    instantS += f"StepN 최저가:  {min_col} USD -{catMin}" + "\n"
    instantS += f"StepN 최고가:  {max_col} USD -{catMax}" + "\n"
    return instantS

def firstGC(you_ = "default", color = False):
    rat_col = fo2(data["market"]["SOLGSTRatio"])
    if color:
        rat_c = "red"
        if data["market"]["SOLGSTRatio"] > user_data[you_]["ratioMax"]:
            rat_c = "green"
        elif data["market"]["SOLGSTRatio"] < user_data[you_]["ratioMin"]:
            rat_c = "blue"        
        rat_col = colored(rat_col, rat_c)
    instantS = ""
    instantS += "\n".join([f"{BTC} FT가:  {fo3(data['market']['base'][BTC]['bid'])} {fo4(data['market']['base'][BTC]['ask'])}" for BTC in data["coinlist"]]) + "\n"
    instantS += f"StepN 원가:  {fo2(data['market']['cost'])} USD" + "\n"
    instantS += f"SOL/GST: {rat_col}" + "\n"
    return instantS

def finalG():
    instantS = f"현재 시각: {time.strftime('%Y-%m-%d-%H:%M:%S')}\n"
    return instantS

def refreshAlertOverlap():
    data["alertOK"] = {you: True for you in info["subscribers"] + ["default"]}    

def dataFormat(obj):
    try:
        inst = f"가격:{obj['price']}\n"
        return inst
    except KeyError:
        return ""

def handle(msg):
    content_type, _, chat_id = telepot.glance(msg)
    chat_id = str(chat_id)
    if content_type == "text":
        msg_params = msg["text"].split(" ")
        command = msg_params[0].lower()
        if chat_id in info["subscribers"]:
            if len(msg_params) > 1:
                content = msg_params[1].replace("\n","")
                if number(content):
                    if command in ["ratiomax", "ratiomin"]:
                        user_data[chat_id][["ratioMax", "ratioMin"][["max", "min"].index(command)]] = float(content)
                        saveFile()
                        bot.sendMessage(chat_id, f"{command}->{content}")
                    elif command in list(manifest["commands_admin"].keys()):
                        if chat_id in info["admin"]:
                            if command in ["ratiomaxd", "ratiomind"]:
                                user_data["default"][["ratioMax", "ratioMin"][["maxd", "mind"].index(command)]] = float(content)
                                saveFile()
                                bot.sendMessage(chat_id, f"{command}->{content}")
                                
                            elif command in ["white_add", "whitelist"]:
                                if content not in info["whitelist"]:
                                    info["whitelist"].append(content)
                                    saveInfo()
                                    bot.sendMessage(chat_id, f"{content} added")
                
                            elif command == "screenshot":
                                dimension = setting["dimension"]
                                try:
                                    if len(msg_params) >= 5:
                                        dimension = tuple(int(c) for c in msg_params[1:5])
                                except IndexError:
                                    pass
                                bot.sendMessage(chat_id, "Processing...")
                                img = pyautogui.screenshot(region = dimension)
                                img_name = f"SC_{time.strftime('%Y%m%d%H%M%S')}.png"
                                img.save("./img/" + img_name)
                                with open("./img/" + img_name, "rb") as fil_:
                                    bot.sendPhoto(chat_id, fil_)
                                os.remove("./img/" + img_name)
                
                            else:
                                bot.sendMessage(chat_id, "None")
                    else:
                        bot.sendMessage(chat_id, "명령어 리스트: " + ", ".join(list(manifest["commands"].keys())) + "\n값 입력은 뒤에 숫자를 넣는다\nex) 환율 40")
                else:
                    if command == "usage":
                        if content in manifest["commands"]:
                            bot.sendMessage(chat_id, f"{content}: {manifest['commands'][content]}")
                        else:
                            bot.sendMessage(chat_id, "\n".join([f"{com}:{manifest['commands'][com]}" for com in manifest["commands"]]))
                    elif command in ["퍼센트", "pricemax", "pricemin"]:
                        if content.lower() in setting["categories"] + ["general"]:
                            if len(msg_params) > 2 and number(msg_params[2]):
                                user_data[chat_id][["percentage", "priceMax", "priceMin"][["퍼센트", "pricemax", "pricemin"].index(command)]][content.lower()] = float(msg_params[2])
                                bot.sendMessage(chat_id, f"{command}/{content}->{msg_params[2]}")
                                saveFile()
                            else:
                                content = user_data[chat_id][["percentage", "priceMax", "priceMin"][["퍼센트", "pricemax", "pricemin"].index(command)]][content.lower()]
                                bot.sendMessage(chat_id, f"{command}/{msg_params[1]}={content}")
                        else:
                            bot.sendMessage(chat_id, "Invalid category name")
                    elif chat_id in info["admin"]:
                        if command in ["퍼센트d", "pricemaxd", "pricemind"]:
                            if content.lower() in setting["categories"] + ["general"]:
                                if len(msg_params) > 2 and number(msg_params[2]):
                                    user_data["default"][["percentage", "priceMax", "priceMin"][["퍼센트d", "pricemaxd", "pricemind"].index(command)]][content.lower()] = float(msg_params[2])
                                    bot.sendMessage(chat_id, f"{command}/{content}->{msg_params[2]}")
                                    saveFile()
                                else:
                                    content = user_data["default"][["percentage", "priceMax", "priceMin"][["퍼센트d", "pricemaxd", "pricemind"].index(command)]][content.lower()]
                                    bot.sendMessage(chat_id, f"{command}/{msg_params[1]}={content}")
                            else:
                                bot.sendMessage(chat_id, "Invalid category name")
                        else:
                            bot.sendMessage(chat_id, "Message not acceptable")
                    else:
                        bot.sendMessage(chat_id, "Message not acceptable")
            else:
                if command in ["ratiomax", "ratiomin"]:
                    content = user_data[chat_id][["ratioMax", "ratioMin"][["max", "min"].index(command)]]
                    bot.sendMessage(chat_id, f"{command}={content}")
                elif command == "데이터":
                    bot.sendMessage(chat_id, str(data))
                elif command == "유저데이터":
                    bot.sendMessage(chat_id, dataFormat(user_data[chat_id]))
                elif command == "기본데이터":
                    bot.sendMessage(chat_id,dataFormat(user_data["default"]))
                elif command == "프린트":
                    instant = firstGS(you_ = chat_id) + "\n"
                    instant += firstGC(you_ = chat_id) + "\n"
                    for category_ in setting["categories"]:
                        instant += printerG(category_, you_ = chat_id)
                    instant += finalG()
                    bot.sendMessage(chat_id, instant)
                elif command == "기본프린트":
                    instant = firstGS() + "\n"
                    instant += firstGC() + "\n"
                    for category_ in setting["categories"]:
                        instant += printerG(category_)
                    instant += finalG()
                    bot.sendMessage(chat_id, instant)
                elif command in ["링크", "초대", "초대링크"]:
                    bot.sendMessage(chat_id, "t.me/" + info["bot_id"])
                elif command == "help":
                    bot.sendMessage(chat_id, "command list: " + ", ".join(list(manifest["commands"].keys())))
                elif command == "usage":
                    bot.sendMessage(chat_id, "\n".join([f"{com}:{manifest['commands'][com]}" for com in manifest["commands"]]))
                elif command == "formula":
                    bot.sendMessage(chat_id, info["formula"])
                elif command == "server":
                    bot.sendMessage(chat_id, fl(info["server_list"], socket.gethostname()))
                elif command in list(manifest["commands_admin"].keys()):
                    if chat_id in info["admin"]:
                        if command == "refresh_code":
                            data["invitation_code"] = randString(7)
                        elif command == "whole_data":
                            whole_lis=[key + "\n"+dataFormat(user_data[key]) for key in user_data.keys()]
                            try:
                                bot.sendMessage(chat_id, "\n".join(whole_lis))
                            except telepot.exception.TelepotException:
                                for x in range(int(len("\n".join(whole_lis)) / 4095) + 1):
                                    bot.sendMessage(chat_id, "\n".join(whole_lis)[x * 4095:(x + 1) * 4095])
                        elif command == "naked_data":
                            try:
                                bot.sendMessage(chat_id, json.dumps(user_data, indent = 4))
                            except telepot.exception.TelepotException:
                                for x in range(int(len(json.dumps(user_data, indent = 4))/4095)+1):
                                    bot.sendMessage(chat_id, json.dumps(user_data, indent = 4)[x * 4095:(x + 1) * 4095])
                        elif command == "code":
                            bot.sendMessage(chat_id, data["invitation_code"])
                        elif command == "us":
                            bot.sendMessage(chat_id, "\n".join([fl(info["subscriber_name_dict"], you) for you in info["subscribers"]]))
                        elif command == "white":
                            bot.sendMessage(chat_id, "\n".join(info["whitelist"]))
                        elif command == "revive":
                            bot.sendMessage(chat_id, "See you soon")
                            data["revive"] = True
                        elif command == "die":
                            bot.sendMessage(chat_id, "See you later")
                            data["die"] = True
                        elif command in ["ratiomaxd", "ratiomind"]:
                            content=user_data["default"][["ratioMax", "ratioMin"][["ratiomaxd", "ratiomind"].index(command)]]
                            bot.sendMessage(chat_id, f"{command}={content}")
                        elif command == "helpd":
                            bot.sendMessage(chat_id, "command list: " + ", ".join(list(manifest["commands_admin"].keys())))
                        elif command == "usaged":
                            bot.sendMessage(chat_id, "\n".join([f"{com}:{manifest['commands_admin'][com]}" for com in manifest["commands_admin"]]))
                        elif command == "screenshot":
                            bot.sendMessage(chat_id, "Processing...")
                            img = pyautogui.screenshot(region = setting["dimension"])
                            img_name = f"S_{time.strftime('%Y%m%d%H%M%S')}.png"
                            img.save("./img/" + img_name)
                            with open("./img/" + img_name, "rb") as fil_:
                                bot.sendPhoto(chat_id, fil_)
                            os.remove("./img/" + img_name)
                        else:
                            bot.sendMessage(chat_id, "None, command list: " + ", ".join(list(manifest["commands_admin"].keys())))
                    else:
                        bot.sendMessage(chat_id, "Forbidden access")
                else:
                    bot.sendMessage(chat_id, "명령어 리스트: " + ", ".join(list(manifest["commands"].keys())) + "\n값 입력은 뒤에 숫자를 넣는다\nex) 환율 40")
        elif chat_id in info["whitelist"]:
            info["subscribers"].append(chat_id)
            saveInfo()
            user_data[chat_id] = user_data["default"]
            data["recent_ban_date"][chat_id] = "0"
            bot.sendMessage(chat_id, "Welcome")
            alert_bot.sendMessage(info["developer"], f"{manifest['name']}: {chat_id} was added")   
        else:
            if command == data["invitation_code"]:
                info["subscribers"].append(chat_id)
                user_data[chat_id] = user_data["default"]
                data["recent_ban_date"][chat_id] = "0"
                bot.sendMessage(chat_id, "Welcome")
                alert_bot.sendMessage(info["developer"], f"{chat_id} was added")          
                data["invitation_code"] = randString(7)
            else:
                bot.sendMessage(chat_id, f"Unknown access. Contact developer t.me/{info['developer_id']} for the key")
                bot.sendMessage(info["developer"], f"{chat_id} tried to access: {msg['text']}")

bot.message_loop(handle)
try:
    while not data["die"]:
        if data["execute_on"] <= time.time() and (time.time() % info["period"] < (info["period"] / 10) or data["initial"]):
            data["cycle_i"] = time.time()
            ws1[f"A{ws1.max_row+1}"] = time.strftime("%Y%m%d%H%M%S")
            getcoin()
            data["prev_data"] = data["market"]
            refreshAlertOverlap()
            
            print("\n\n\n\n\n\n\n\n\n\n\n\n")
            print(firstGS(you_ = "default", color = True))
            print(firstGC(you_ = "default", color = True))
            for category in setting["categories"]:
                print(printerG(category, you_ = "default", color = True))
            print(finalG())
            
            premiums = [data["market"]["shoes"][category]["premium"] for category in setting["categories"]]
            prices = [data["market"]["shoes"][category]["priceUSD"] for category in setting["categories"]]
            maxPreCat = setting["categories"][premiums.index(max(premiums))]
            maxPriCat = setting["categories"][prices.index(max(prices))]
            minPriCat = setting["categories"][prices.index(min(prices))]
            
            if data["market"]["shoes"][maxPreCat]["premium"] > user_data["default"]["percentage"]["general"] and data["alertOK"]["default"] and data["market"]["shoes"][maxPreCat]["valid"]:
                data["alertOK"]["default"] = False
            elif data["market"]["shoes"][maxPriCat]["priceUSD"] > user_data["default"]["priceMax"]["general"] and data["alertOK"]["default"] and data["market"]["shoes"][maxPriCat]["valid"]:
                data["alertOK"]["default"] = False
            elif data["market"]["shoes"][minPriCat]["priceUSD"] < user_data["default"]["priceMin"]["general"] and data["alertOK"]["default"] and data["market"]["shoes"][minPriCat]["valid"]:
                data["alertOK"]["default"] = False
            for category in setting["categories"]:
                if data["market"]["shoes"][category]["premium"] > user_data["default"]["percentage"][category] and data["alertOK"]["default"] and data["market"]["shoes"][category]["valid"]:
                    data["alertOK"]["default"] = False
                elif data["market"]["shoes"][category]["priceUSD"] > user_data["default"]["priceMax"][category] and data["alertOK"]["default"] and data["market"]["shoes"][category]["valid"]:
                    data["alertOK"]["default"] = False
                elif data["market"]["shoes"][category]["priceUSD"] < user_data["default"]["priceMin"][category] and data["alertOK"]["default"] and data["market"]["shoes"][category]["valid"]:
                    data["alertOK"]["default"] = False
            if data["market"]["SOLGSTRatio"] > user_data["default"]["ratioMax"] and data["alertOK"]["default"]:
                data["alertOK"]["default"] = False
            elif data["market"]["SOLGSTRatio"] < user_data["default"]["ratioMin"] and data["alertOK"]["default"]:
                data["alertOK"]["default"] = False
            for you in info["subscribers"]:
                i = 0
                while i < info["msgTry"]:
                    try:
                        if data["market"]["shoes"][maxPreCat]["premium"] > user_data[you]["percentage"]["general"] and data["alertOK"][you] and data["market"]["shoes"][maxPreCat]["valid"]:
                            bot.sendMessage(you, firstGC() + "\n" + printerG(maxPreCat, you_ = you))
                            data["alertOK"][you] = False
                        elif data["market"]["shoes"][maxPriCat]["priceUSD"] > user_data[you]["priceMax"]["general"] and data["alertOK"][you] and data["market"]["shoes"][maxPriCat]["valid"]:
                            bot.sendMessage(you, firstGC() + "\n" + printerG(maxPriCat, you_ = you))
                            data["alertOK"][you] = False
                        elif data["market"]["shoes"][minPriCat]["priceUSD"] < user_data[you]["priceMin"]["general"] and data["alertOK"][you] and data["market"]["shoes"][minPriCat]["valid"]:
                            bot.sendMessage(you, firstGC() + "\n" + printerG(minPriCat, you_ = you))
                            data["alertOK"][you] = False
                        for category in setting["categories"]:
                            if data["market"]["shoes"][category]["premium"] > user_data[you]["percentage"][category] and data["alertOK"][you] and data["market"]["shoes"][category]["valid"]:
                                bot.sendMessage(you, firstGC() + "\n" + printerG(category, you_ = you))
                                data["alertOK"][you] = False
                            elif data["market"]["shoes"][category]["priceUSD"] > user_data[you]["priceMax"][category] and data["alertOK"][you] and data["market"]["shoes"][category]["valid"]:
                                bot.sendMessage(you, firstGC() + "\n" + printerG(category, you_ = you))
                                data["alertOK"][you] = False
                            elif data["market"]["shoes"][category]["priceUSD"] < user_data[you]["priceMin"][category] and data["alertOK"][you] and data["market"]["shoes"][category]["valid"]:
                                bot.sendMessage(you, firstGC() + "\n" + printerG(category, you_ = you))
                                data["alertOK"][you] = False
                        if data["market"]["SOLGSTRatio"] > user_data[you]["ratioMax"] and data["alertOK"][you]:
                            bot.sendMessage(you, firstGC())
                            data["alertOK"][you] = False
                        elif data["market"]["SOLGSTRatio"] < user_data[you]["ratioMin"] and data["alertOK"][you]:
                            bot.sendMessage(you, firstGC())
                            data["alertOK"][you] = False
                        if not data["alertOK"][you]:
                            time.sleep(info["msgGap"])
                        break
                    except telepot.exception.TelepotException:
                        if data["recent_ban_date"][you] != time.strftime("%Y-%m-%d"):
                            alert_bot.sendMessage(info["developer"], f"{manifest['name']}: {fl(info['subscriber_name_dict'], you)} blocked this bot")
                            data["recent_ban_date"][you] = time.strftime("%Y-%m-%d")
                    i += 1
            data["cycle_f"] = time.time()
            if not data["initial"]:
                time.sleep(max(0, info["period"] / 10 - (time.time() - data["cycle_i"])))
            else:
                data["initial"] = False
        else:
            time.sleep(info["period"] / 15)
        if data["revive"]:
            sys.exit()
        
except:
    data["error"] = traceback.format_exc()
    os.system("start cmd /C " + sys.argv[0])
    winsound.PlaySound("./data/에러.wav", winsound.SND_ALIAS)

print(f"error {time.strftime('%Y-%m-%d-%H:%M:%S')}")
wb.save(f"./dataxl/{manifest['name']}{data['start']}.xlsx")
with open("./log/" + sys.argv[0].split('\\')[-1].split('.')[0] + ".txt", "a", encoding = "utf-8") as fil:
    fil.write(f"시간: {time.strftime('%Y-%m-%d-%H:%M:%S')}\n\n{data['error']}\n---------------------------\n\n")